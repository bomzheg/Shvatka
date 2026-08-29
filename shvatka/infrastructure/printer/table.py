import typing
from dataclasses import dataclass
from io import BytesIO

from openpyxl import Workbook
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.chart.axis import ChartLines
from openpyxl.chart.marker import Marker
from openpyxl.chart.series import Series
from openpyxl.chart.series_factory import SeriesFactory
from openpyxl.chart.shapes import GraphicalProperties
from openpyxl.drawing.colors import ColorChoice
from openpyxl.drawing.fill import PatternFillProperties
from openpyxl.drawing.line import LineProperties
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from shvatka.core.interfaces.printer import (
    CellStyle,
    Chart,
    SeriesKind,
    Table,
    TablePrinter,
)

FONT_NAME = "Verdana"
INK = "FF222222"
BEST_INK = "FFC9211E"
HEADER_FILL = "FFD6DCE5"
TEAM_FILL = "FFDFE6EF"
GRID = "FFBFBFBF"
SURFACE = "FFFFFF"

GENERAL_FORMAT = "General"
MAX_COLUMN_WIDTH = 32
MIN_COLUMN_WIDTH = 4
COLUMN_PADDING = 1
DATA_FONT_SCALE = 0.85
"""Column width is counted in the default font; ours is smaller than that."""

SERIES_COLORS = (
    "2A78D6",
    "EB6834",
    "1BAF7A",
    "EDA100",
    "E87BA4",
    "008300",
    "4A3AA7",
    "E34948",
)
"""Categorical palette, assigned in this fixed order — never re-ordered per game."""

ACCENT_COLOR = "52514E"
"""Reference series (an average) — neutral ink, so it doesn't read as one more team."""

SeriesPattern = typing.Literal["ltUpDiag", "ltHorz"]
SERIES_PATTERNS: tuple[SeriesPattern | None, ...] = (None, "ltUpDiag", "ltHorz")
"""Texture per lap over the palette, so a ninth team is told apart from the first."""

HAIRLINE_WIDTH = 6350
"""EMU — half a point."""
LINE_WIDTH = 25400
"""EMU — two points."""


@dataclass(frozen=True)
class CellFormat:
    font: Font | None = None
    fill: PatternFill | None = None
    border: Border | None = None
    alignment: Alignment | None = None


def _side() -> Side:
    return Side(style="thin", color=GRID)


BOXED = Border(left=_side(), right=_side(), top=_side(), bottom=_side())

FORMATS: dict[CellStyle, CellFormat] = {
    CellStyle.PLAIN: CellFormat(font=Font(name=FONT_NAME, size=9, color=INK)),
    CellStyle.TITLE: CellFormat(font=Font(name=FONT_NAME, size=14, bold=True, color=INK)),
    CellStyle.SECTION: CellFormat(font=Font(name=FONT_NAME, size=10, bold=True, color=INK)),
    CellStyle.HEADER: CellFormat(
        font=Font(name=FONT_NAME, size=8, bold=True, color=INK),
        fill=PatternFill("solid", fgColor=HEADER_FILL),
        border=BOXED,
        alignment=Alignment(horizontal="center", vertical="center", wrap_text=True),
    ),
    CellStyle.TEAM: CellFormat(
        font=Font(name=FONT_NAME, size=8, bold=True, color=INK),
        fill=PatternFill("solid", fgColor=TEAM_FILL),
        border=BOXED,
        alignment=Alignment(vertical="center", wrap_text=True),
    ),
    CellStyle.DATA: CellFormat(
        font=Font(name=FONT_NAME, size=8, color=INK),
        border=BOXED,
        alignment=Alignment(horizontal="center", vertical="center", wrap_text=True),
    ),
    CellStyle.ACCENT: CellFormat(
        font=Font(name=FONT_NAME, size=8, bold=True, color=INK),
        fill=PatternFill("solid", fgColor=HEADER_FILL),
        border=BOXED,
        alignment=Alignment(horizontal="center", vertical="center"),
    ),
    CellStyle.BEST: CellFormat(
        font=Font(name=FONT_NAME, size=8, bold=True, color=BEST_INK),
        border=BOXED,
        alignment=Alignment(horizontal="center", vertical="center"),
    ),
}

WIDTH_STYLES = frozenset(
    {CellStyle.TEAM, CellStyle.DATA, CellStyle.ACCENT, CellStyle.BEST, CellStyle.PLAIN}
)
"""Styles a column is sized to fit. Headers and captions wrap instead."""


class ExcellPrinter(TablePrinter):
    def print_table(self, table: Table) -> BytesIO:
        result = BytesIO()
        print_table(table, result)
        result.seek(0)
        return result


def print_table(table: Table, file: typing.Any) -> None:
    wb = Workbook()
    ws = typing.cast(Worksheet, wb.active)
    ws.sheet_view.showGridLines = False
    widths: dict[int, int] = {}
    for address, cell_ in table.fields.items():
        cell = ws.cell(**address.to_dict())
        cell.value = cell_.value
        if cell_.format is not None:
            cell.number_format = cell_.format
        apply_style(cell, cell_.style)
        if cell_.style in WIDTH_STYLES:
            widths[address.column] = max(widths.get(address.column, 0), _display_len(cell))
    resize_columns(ws, widths)
    for column in table.hidden_columns:
        ws.column_dimensions[get_column_letter(column)].hidden = True
    if table.freeze is not None:
        ws.freeze_panes = ws.cell(**table.freeze.to_dict())
    for chart in table.charts:
        add_chart(ws, chart)
    wb.save(file)


def apply_style(cell: typing.Any, style: CellStyle) -> None:
    format_ = FORMATS[style]
    if format_.font is not None:
        cell.font = format_.font
    if format_.fill is not None:
        cell.fill = format_.fill
    if format_.border is not None:
        cell.border = format_.border
    if format_.alignment is not None:
        cell.alignment = format_.alignment


def resize_columns(worksheet: Worksheet, widths: dict[int, int]) -> None:
    """Fit every column to its data. Headers are wrapped rather than fitted — a long
    level name would otherwise stretch a column of eight-character times.
    """
    for column, width in widths.items():
        worksheet.column_dimensions[get_column_letter(column)].width = min(
            max(width * DATA_FONT_SCALE + COLUMN_PADDING, MIN_COLUMN_WIDTH), MAX_COLUMN_WIDTH
        )


def _display_len(cell: typing.Any) -> int:
    """Width the cell takes on screen — what is shown, not what is stored.

    A datetime stored behind ``HH:MM:SS`` shows eight characters, not the
    nineteen its ``str()`` has, so the number format is the better measure.
    """
    if cell.value is None:
        return 0
    if cell.number_format != GENERAL_FORMAT:
        return len(cell.number_format)
    return max(len(line) for line in str(cell.value).splitlines() or [""])


def add_chart(worksheet: Worksheet, chart: Chart) -> None:
    bars = BarChart()
    bars.type = "col"
    bars.grouping = "clustered"
    bars.gapWidth = 60
    lines = LineChart()
    for i, series in enumerate(chart.series):
        built = SeriesFactory(
            Reference(
                worksheet,
                min_col=series.title.column,
                min_row=series.title.row,
                max_col=series.value_range.end.column,
                max_row=series.value_range.end.row,
            ),
            title_from_data=True,
        )
        target = lines if series.kind is SeriesKind.LINE else bars
        _paint(built, i, series.accent, series.kind)
        target.series.append(built)
    categories = Reference(
        worksheet,
        min_col=chart.categories.start.column,
        min_row=chart.categories.start.row,
        max_col=chart.categories.end.column,
        max_row=chart.categories.end.row,
    )
    bars.set_categories(categories)
    lines.set_categories(categories)
    if lines.series:
        bars += lines
    _decorate(bars, chart)
    worksheet.add_chart(bars, _anchor(chart))


def _anchor(chart: Chart) -> str:
    return f"{get_column_letter(chart.anchor.column)}{chart.anchor.row}"


def _decorate(built: BarChart, chart: Chart) -> None:
    built.title = chart.title
    built.style = 2
    built.width = chart.width
    built.height = chart.height
    built.x_axis.title = chart.x_title
    built.y_axis.title = chart.y_title
    if chart.y_format is not None:
        built.y_axis.number_format = chart.y_format
    built.y_axis.majorGridlines = ChartLines(spPr=_hairline())
    built.x_axis.spPr = _hairline()
    built.legend.position = "b"
    built.legend.overlay = False


def _paint(series: Series, index: int, accent: bool, kind: SeriesKind) -> None:
    if accent:
        color, pattern = ACCENT_COLOR, None
    else:
        color = SERIES_COLORS[index % len(SERIES_COLORS)]
        pattern = SERIES_PATTERNS[(index // len(SERIES_COLORS)) % len(SERIES_PATTERNS)]
    properties = GraphicalProperties()
    if kind is SeriesKind.LINE:
        properties.line = LineProperties(solidFill=color, w=LINE_WIDTH)
        series.marker = Marker(symbol="none")
        series.smooth = False
    else:
        properties.line = LineProperties(noFill=True)
        if pattern is None:
            properties.solidFill = color
        else:
            properties.pattFill = PatternFillProperties(
                prst=pattern, fgClr=_color(color), bgClr=_color(SURFACE)
            )
    series.graphicalProperties = properties


def _hairline() -> GraphicalProperties:
    return GraphicalProperties(ln=LineProperties(solidFill=GRID[2:], w=HAIRLINE_WIDTH))


def _color(rgb: str) -> ColorChoice:
    """openpyxl takes a plain hex string here, whatever the type stubs promise."""
    color = ColorChoice()
    color.RGB = rgb
    return color
