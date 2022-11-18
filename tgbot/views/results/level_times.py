import typing
from dataclasses import dataclass
from datetime import datetime

from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from shvatka.models import dto
from shvatka.utils.exceptions import GameNotFinished


@dataclass
class CellAddress:
    column: int
    row: int

    def shift(self, plus_rows=0, plus_columns=0):
        return dict(row=self.row + plus_rows, column=self.column + plus_columns)


FIRST_TEAM_NAME = CellAddress(1, 3)
GAME_NAME = CellAddress(1, 1)


class LevelTime(typing.NamedTuple):
    level: int
    time: datetime


class TeamLevelsTimes(typing.NamedTuple):
    team: dto.Team
    levels_times: list[LevelTime]


DATETIME_EXCEL_FORMAT = "HH:MM:SS"


def export_results(game: dto.FullGame, game_stat: dto.GameStat, file: typing.Any):
    return export_results_internal(game, to_array_results(game_stat), file)


def to_array_results(game_stat: dto.GameStat) -> list[TeamLevelsTimes]:
    result = []
    for team, lts in game_stat.level_times.items():
        levels_times = [LevelTime(lt.level_number, lt.start_at) for lt in lts]
        result.append(TeamLevelsTimes(team, levels_times))
    return result


def export_results_internal(game: dto.FullGame, data: list[TeamLevelsTimes], file: typing.Any):
    if not (game.is_complete() or game.is_finished()):
        raise GameNotFinished
    wb = Workbook()
    ws = wb.active
    ws.cell(**GAME_NAME.shift()).value = game.name
    for i, team_level_times in enumerate(data):
        cell = ws.cell(**FIRST_TEAM_NAME.shift(i, 0))
        cell.value = team_level_times.team.name

        for j, level_time in enumerate(team_level_times.levels_times, 1):
            if i == 0:
                ws.cell(**FIRST_TEAM_NAME.shift(-1, j)).value = level_time.level

            cell = ws.cell(**FIRST_TEAM_NAME.shift(i, j))
            cell.value = level_time.time
            cell.number_format = DATETIME_EXCEL_FORMAT

    resize_columns(ws)
    wb.save(file)


def resize_columns(worksheet: Worksheet):
    for col in worksheet.columns:
        new_len = max([2, *[len(str(cell.value or "")) for cell in col]])
        worksheet.column_dimensions[col[0].column_letter].width = new_len
